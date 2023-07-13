from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, PatternFill
import urllib.request
import os
import tkinter as tk
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class Order:
    def __init__(self, name, street, apartment, city, country, image, size, final, date, isStarted):
        self.name = name
        self.street = street
        self.apartment = apartment
        self.city = city
        self.country = country
        self.image = image
        self.size = size
        self.final = final
        self.date = date
        self.isStarred = isStarted

class DateObj:
    def __init__(self, day, month, year):
        self.day = day
        self.month = month
        self.year = year


def whatMonth(month):
    # month = month.capitalize()
    if month == "Jun":
        return 1
    elif month == "Feb":
        return 2
    elif month == "Mar":
        return 3
    elif month == "Apr":
        return 4
    elif month == "May":
        return 5
    elif month == "Jun":
        return 6
    elif month == "Jul":
        return 7
    elif month == "Aug":
        return 8
    elif month == "Sep":
        return 9
    elif month == "Oct":
        return 10
    elif month == "Nov":
        return 11
    elif month == "Dec":
        return 12
    elif month == "1":
        return "Jan"
    elif month == "2":
        return "Feb"
    elif month == "3":
        return "Mar"
    elif month == "4":
        return "Apr"
    elif month == "5":
        return "May"
    elif month == "6":
        return "Jun"
    elif month == "7":
        return "Jul"
    elif month == "8":
        return "Aug"
    elif month == "9":
        return "Sep"
    elif month == "10":
        return "Oct"
    elif month == "11":
        return "Nov"
    elif month == "12":
        return "Dec"
    

# USER INTERFACE
def submit():
    dateObj.day = day_entry.get()
    dateObj.month = whatMonth(month_entry.get())
    dateObj.year = year_entry.get()

    root.quit()

root = tk.Tk()
root.title("Etsy Bot")
root.geometry("700x350")

day_label = tk.Label(root, text="Gün: (örneğin 1, 13, 21, ...))")
day_label.pack()

day_entry = tk.Entry(root)
day_entry.pack()

month_label = tk.Label(root, text="Ay: (örneğin 1, 13, 21, ...)")
month_label.pack()

month_entry = tk.Entry(root)
month_entry.pack()

year_label = tk.Label(root, text="Yıl: (örneğin 2022, 2023, ...)")
year_label.pack()

year_entry = tk.Entry(root)
year_entry.pack()

dateObj = DateObj("", "", "")
canStart = False
submit_button = tk.Button(root, text="Başlat", command=submit)
submit_button.pack()

root.mainloop()
root.destroy()

stringToday = str(whatMonth(dateObj.month)) + "." + dateObj.day + "." + dateObj.year


# New Director Path
if not os.path.exists(stringToday):
    os.mkdir(stringToday)
    folderPath = stringToday
    os.mkdir(f"./{folderPath}/{stringToday} fotograflar")
    os.mkdir(f"./{folderPath}/{stringToday} excel")
    os.mkdir(f"./{folderPath}/{stringToday} excel/fotograflar")
else:
    z = 1
    while True:
        if not os.path.exists(stringToday + " (" + str(z) + ")"):
            os.mkdir(stringToday + " (" + str(z) + ")")
            folderPath = stringToday + " (" + str(z) + ")"
            os.mkdir(f"./{folderPath}/{stringToday} fotograflar")
            os.mkdir(f"./{folderPath}/{stringToday} excel")
            os.mkdir(f"./{folderPath}/{stringToday} excel/fotograflar")
            
            break
        z += 1
# ************************************************


PATH = "C:\Program Files (x86)\msedgedriver.exe"
driver = webdriver.Edge()
wait = WebDriverWait(driver, 15)

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.column_dimensions['A'].width = 11
sheet.column_dimensions['B'].width = 11
sheet.column_dimensions['C'].width = 11
sheet.column_dimensions['D'].width = 11
sheet.column_dimensions['E'].width = 40
sheet.column_dimensions['F'].width = 30
workbook.save(f"./{folderPath}/{stringToday} excel/{stringToday}.xlsx")



driver.get("https://www.etsy.com/")
driver.maximize_window()

signIn = driver.find_element(
    By.XPATH, "/html/body/div[2]/header/div[4]/nav/ul/li[1]/button")
signIn.click()
wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='join_neu_email_field']")))

emailInput = driver.find_element(By.XPATH, "//*[@id='join_neu_email_field']")
emailInput.send_keys("estveryin15@gmail.com")

passwordInput = driver.find_element(
    By.XPATH, "//*[@id='join_neu_password_field']")
passwordInput.send_keys("2022Th3005")

signInButton = driver.find_element(
    By.XPATH, "/html/body/div[5]/div[2]/div/div[3]/div/div/div/div/div/div[2]/form/div[1]/div/div[7]/div/button")
signInButton.click()
time.sleep(5)

try:
    driver.find_element(By.XPATH, "//*[@id='join-neu-overlay']")
    time.sleep(20)
except:
    wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))


shopManager = driver.find_element(
    By.XPATH, "//*[@id='gnav-header-inner']/div[4]/nav/ul/li[3]")
shopManager.click()
wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))


od = driver.find_element(
    By.XPATH, "//*[@id='root']/div/div[1]/div[3]/div/div[1]/div[2]/ul/li[5]")
od.click()
time.sleep(8)

def start(cell_row_N):
    cell_row = cell_row_N
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    k = 2
    while True:
        dispatchXP = "/html/body/div[3]/div/div[1]/div/main/div/div[2]/div/div/div[1]/div[3]/div[" + str(k) + "]"

        try:
            dispatch = driver.find_element(By.XPATH, dispatchXP)
        except:
            break

        i = 1
        while True:
            xpath = dispatchXP + "/div[2]/div[" + str(i) + "]"

            try:
                data = driver.find_element(By.XPATH, xpath)
                order = Order("", "", "", "", "", [], [], [], str(whatMonth(dateObj.month)) + "/" + dateObj.day + "/" + dateObj.year, False)
                workbook = openpyxl.load_workbook(f"./{folderPath}/{stringToday} excel/{stringToday}.xlsx")
                sheet = workbook.active
            except:
                break

            dateOfOrderPath = xpath + "/div/div[2]/div/div[2]/div[2]"
            dateList = driver.find_element(By.XPATH, dateOfOrderPath).text.split(" ")
            dateOfOrder = str(whatMonth(dateList[2].split(",")[0])) + "/" + dateList[1] + "/" + dateList[3]
            if not (dateOfOrder == order.date):
                i += 1
                continue


            try:
                repeatBuyer = driver.find_element(By.XPATH, xpath + "/div/div[2]/div/div[1]/div[1]/div[1]/div[1]/div/span[2]/span[1]/span[2]")
                if repeatBuyer.text == "Repeat buyer":
                    order.isStarred = True
            except:
                order.isStarred = False

            # TASK 2*************************************************************************************************
            driver.find_element(By.XPATH, xpath + "/div/div[2]/div/div[2]/div[1]").click()
            a = 2
            original_window = driver.current_window_handle
            while True:
                linkPath = "/html/body/div[3]/div/div[1]/div/main/div/div[3]/div[2]/div/div/div[9]/div/div/div[2]/table/tbody/tr[" + str(a)  + "]/td[1]/div[2]/div[2]/a"

                try:
                    # wait.until(EC.element_to_be_clickable((By.XPATH, linkPath)))
                    link = driver.find_element(By.XPATH, linkPath)
                except:
                    try:
                        linkPath = "/html/body/div[3]/div/div[1]/div/main/div/div[3]/div[2]/div/div/div[10]/div/div/div[2]/table/tbody/tr[" + str(a)  + "]/td[1]/div[2]/div[2]/a"
                        link = driver.find_element(By.XPATH, linkPath)
                    except:
                        if a == 2:
                            print(f"Fotoğraf kaydedilemedi: Baştan {i}. ürün")
                        break
                
                link.click()
                time.sleep(0.15)
                driver.switch_to.window(driver.window_handles[1])

                wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                
                try:
                    imgSrc = driver.find_element(By.XPATH, "/html/body/main/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div[1]/ul/li[1]/img").get_attribute("src")
                except:
                    imgSrc = driver.find_element(By.XPATH, "//*[@id='content']/div/div[1]/div/div/div[2]/div[1]/div/img")
                
                try:
                    fileName = driver.find_element(By.XPATH, "/html/body/main/div[1]/div[2]/div/div/div[1]/div[2]/div/div[4]/h1").text
                except:
                    fileName = driver.find_element(By.XPATH, "/html/body/main/div[1]/div[2]/div/div/div[1]/div[2]/div/div[3]/h1").text

                try:
                    r = 1
                    while True:
                        if not os.path.exists(f"./{folderPath}/{stringToday} fotograflar/{fileName}.jpg"):
                            urllib.request.urlretrieve(imgSrc, f"./{folderPath}/{stringToday} fotograflar/{fileName}.jpg")
                            break
                        else:
                            fileName = fileName + " (" + str(r) + ")"
                            r += 1
                except:
                    print("")
                    print(fileName, "indirilemedi")
                    print("")


                driver.close()
                driver.switch_to.window(original_window)
                a += 1
            
            webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            # *************************************************************************************************

            wait.until(EC.element_to_be_clickable((By.XPATH, xpath + "/div/div[2]/div[1]/div[2]/div[4]/div/div[1]/div/button")))
            delivertoBtn = driver.find_element(
                By.XPATH, xpath + "/div/div[2]/div[1]/div[2]/div[4]/div/div[1]/div/button")
            delivertoBtn.click()

            obj = driver.find_element(
                By.XPATH, xpath + "/div/div[2]/div/div[2]/div[4]/div/div[2]/div/div/p[1]")
            id = obj.text.split("\n")
            order.name = id[0]
            order.street = id[1]
            order.apartment = id[2]
            order.city = id[3]
            try:
                order.country = id[4]
            except:
                order.country = ""

            j = 1
            while True:
                isExistPath = xpath + "/div/div[2]/div[2]"
                try:
                    commentLine = driver.find_element(By.XPATH, isExistPath)
                    productPath = xpath + \
                        "/div/div[2]/div[1]/div[1]/div[3]/div/div[" + str(j) + "]"
                except:
                    productPath = xpath + \
                        "/div/div[2]/div/div[1]/div[2]/div/div[" + str(j) + "]"

                try:
                    pro = driver.find_element(By.XPATH, productPath)
                except:
                    break



                try:
                    finishPath = productPath + "/div/div[2]/ul/ul/li[4]/div[1]/span[2]"
                    finishName = driver.find_element(By.XPATH, finishPath)

                    sizePath = productPath + "/div/div[2]/ul/ul/li[3]/div[1]/span[2]"
                    size = driver.find_element(By.XPATH, sizePath)
                    if len(size.text.split(" ")) > 1:
                        order.size.append(size.text)
                    else:
                        order.size.append("")
                except:
                    try:
                        finishPath = productPath + "/div/div[2]/ul/ul/li[3]/div[1]/span[2]"
                        finishName = driver.find_element(By.XPATH, finishPath)
                        
                        sizePath = productPath + "/div/div[2]/ul/ul/li[2]/div[1]/span[2]"
                        size = driver.find_element(By.XPATH, sizePath)
                        if len(size.text.split(" ")) > 1:
                            order.size.append(size.text)
                        else:
                            order.size.append("")
                    except:
                        finishPath = productPath + "/div/div[2]/ul/ul/li[2]/div[1]/span[2]"
                        finishName = driver.find_element(By.XPATH, finishPath)
                        order.size.append("")

                order.final.append(finishName.text)


                img = driver.find_element(
                    By.XPATH, productPath + "/div/div[1]/a/div/div[1]/img")
                src = img.get_attribute("src")
                

                order.image.append(src)

                print()

                j += 1

            l = 0
            for img in order.image:
                imgUrl = f"./{folderPath}/{stringToday} excel/fotograflar/{order.name}.jpg"
                c = 1
                while True:
                    if not os.path.exists(imgUrl):
                        urllib.request.urlretrieve(img, imgUrl)
                        break
                    else:
                        imgUrl = f"./{folderPath}/{stringToday} excel/fotograflar/{order.name}({str(c)}).jpg"
                        c += 1
                
                image = Image(imgUrl)
                sheet.add_image(image, f"{alphabet[l]}{str(cell_row + 1)}")

                sheet[f"{alphabet[l]}{str(cell_row + 5)}"] = order.size[l] + " " + order.final[l]
                sheet[f"{alphabet[l]}{str(cell_row + 5)}"].font = Font(size=8, name="Arial Tur")
                l += 1
            
            sheet[f"A{cell_row}"] = order.date
            sheet[f"C{cell_row}"] = order.date
            sheet[f"A{cell_row}"].font = Font(size=8, name="Arial Tur")
            sheet[f"C{cell_row}"].font = Font(size=8, name="Arial Tur")

            sheet[f"F{cell_row + 1}"] = "Tugce SONUC"
            sheet[f"F{cell_row + 1}"].font = Font(size=11, name="Arial Tur")

            sheet[f"F{cell_row + 2}"] = "Gaziosmanpasa Bulv."
            sheet[f"F{cell_row + 2}"].font = Font(size=11, name="Arial Tur")

            sheet[f"F{cell_row + 3}"] = "No:64 D:601 K:6 Yüncü İşhanı"
            sheet[f"F{cell_row + 3}"].font = Font(size=10, name="Arial Tur")

            sheet[f"F{cell_row + 4}"] = "Cankaya - İzmir"
            sheet[f"F{cell_row + 4}"].font = Font(size=11, name="Arial Tur")



            sheet[f"E{cell_row + 1}"] = "Ship To"
            sheet[f"E{cell_row + 1}"].font = Font(size=11, name="Arial Tur")

            sheet[f"E{cell_row + 2}"] = order.name
            sheet[f"E{cell_row + 2}"].font = Font(size=9, name="Arial Tur", bold=True)

            sheet[f"E{cell_row + 3}"] = order.street
            sheet[f"E{cell_row + 3}"].font = Font(size=9, name="Tahoma")

            sheet[f"E{cell_row + 4}"] = order.apartment
            sheet[f"E{cell_row + 4}"].font = Font(size=9, name="Arial Tur")

            sheet[f"E{cell_row + 5}"] = order.city
            sheet[f"E{cell_row + 5}"].font = Font(size=9, name="Arial Tur")
            
            
            borderBottom = Border(bottom=Side(border_style='medium', color='000000'))
            borderRight = Border(right=Side(border_style='medium', color='000000'))
            borderLeft = Border(left=Side(border_style='medium', color='000000'))
            borderTop = Border(top=Side(border_style='medium', color='000000'))
            
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            r = 1
            for row in sheet[f"A{cell_row}:F{cell_row}"]:
                c = 1
                for cell in row:
                    leftB = c == 1
                    rightB = c == len(row)
                    topB = r == 1
                    bottomB = r == len(sheet[f"A{cell_row}:F{cell_row}"])
                    
                    if leftB:
                        cell.border += borderLeft
                    if rightB:
                        cell.border += borderRight
                    if topB:
                        cell.border += borderTop
                    if bottomB:
                        cell.border += borderBottom

                    c += 1
                r += 1

            if len(order.country) != 0:
                sheet[f"E{cell_row + 6}"] = order.country
                sheet[f"E{cell_row + 6}"].font = Font(size=9, name="Arial Tur")

                sheet[f"F{cell_row + 6}"] = "35240   Turkey"
                sheet[f"F{cell_row + 6}"].font = Font(size=11, name="Arial Tur")

                # BORDERS ***********************************************************
                r = 1
                for row in sheet[f"A{cell_row}:F{cell_row + 6}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"A{cell_row}:D{cell_row + 6}"])

                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom

                        c += 1
                    r += 1

                r = 1
                for row in sheet[f"E{cell_row}:E{cell_row + 6}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"E{cell_row}:E{cell_row + 6}"])
                        
                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom
        
                        c += 1
                    r += 1
                
                r = 1
                for row in sheet[f"F{cell_row}:F{cell_row + 6}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"F{cell_row}:F{cell_row + 6}"])
                        
                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom
        
                        c += 1
                    r += 1
                # ********************************************************************

                # Color **************************************************************
                if order.isStarred == True:
                    for row in sheet[f"A{cell_row}:F{cell_row + 6}"]:
                        for cell in row:
                            cell.fill = fill
                # ********************************************************************

                cell_row += 7
            else:
                sheet[f"F{cell_row + 5}"] = "35240   Turkey"
                sheet[f"F{cell_row + 5}"].font = Font(size=11, name="Arial Tur")

                # BORDERS ***********************************************************
                r = 1
                for row in sheet[f"A{cell_row}:F{cell_row + 5}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"A{cell_row}:D{cell_row + 5}"])

                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom

                        c += 1
                    r += 1

                r = 1
                for row in sheet[f"E{cell_row}:E{cell_row + 5}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"E{cell_row}:E{cell_row + 5}"])
                        
                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom
        
                        c += 1
                    r += 1
                
                r = 1
                for row in sheet[f"F{cell_row}:F{cell_row + 5}"]:
                    c = 1
                    for cell in row:
                        leftB = c == 1
                        rightB = c == len(row)
                        topB = r == 1
                        bottomB = r == len(sheet[f"F{cell_row}:F{cell_row + 5}"])
                        
                        if leftB:
                            cell.border += borderLeft
                        if rightB:
                            cell.border += borderRight
                        if topB:
                            cell.border += borderTop
                        if bottomB:
                            cell.border += borderBottom
        
                        c += 1
                    r += 1
                # ********************************************************************

                # COLOR **************************************************************
                if order.isStarred == True:
                    for row in sheet[f"A{cell_row}:F{cell_row + 5}"]:
                        for cell in row:
                            cell.fill = fill
                # ********************************************************************

                cell_row += 6


            
            i += 1

            workbook.save(f"./{folderPath}/{stringToday} excel/{stringToday}.xlsx")
        k += 1

    return cell_row


cell_r = start(1)

try:
    content = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/div/main/div/div[2]/div/div/div[1]/div[3]")
    navigator = content.find_element(By.XPATH, "(//div)[last()]")
    
    num = 2
    while True:
        try:
            content = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/div/main/div/div[2]/div/div/div[1]/div[3]")
            navigator = content.find_element(By.XPATH, "(//div)[last()]")
            navigator.find_element(By.XPATH, f"//*[text()='{num}']").click()
            num += 1
            cell_r = start(cell_r)

        except:
            break

except:
    print("navigator no exist")


time.sleep(1)
driver.quit()
